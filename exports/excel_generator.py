"""
Gerador de relatórios Excel formatados.

Produz um Excel com formatação profissional (moeda, percentuais, bordas,
estilos) com múltiplas abas para o parecer do Comitê de Crédito.
"""

import io

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from config import theme as t
from domain.pricing import ResultadoPrecificacao
from domain.scoring import Classificacao
from utils.formatters import formatar_cnpj


# =============================================================================
# Estilos reutilizáveis — Institutional Research
# =============================================================================
INK_FILL = PatternFill(start_color="0A0A0B", end_color="0A0A0B", fill_type="solid")
ACCENT_FILL = PatternFill(start_color="8B6A1A", end_color="8B6A1A", fill_type="solid")
ALT_FILL = PatternFill(start_color="FAFAF9", end_color="FAFAF9", fill_type="solid")

# Aliases legados
NAVY_FILL = INK_FILL
GOLD_FILL = ACCENT_FILL

WHITE_BOLD = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
INK_BOLD = Font(name="Calibri", size=11, bold=True, color="0A0A0B")
NAVY_BOLD = INK_BOLD
DEFAULT_FONT = Font(name="Calibri", size=10, color="0A0A0B")

BORDER = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", indent=1)


def gerar_excel_parecer(
    cnpj: str,
    data_hora: str,
    valor: float,
    resultado: ResultadoPrecificacao,
    classificacao: Classificacao,
) -> bytes:
    """Gera o parecer em Excel formatado.

    Returns:
        Bytes do arquivo .xlsx pronto para download.
    """
    output = io.BytesIO()

    # =========================================================================
    # Dados estruturados
    # =========================================================================
    df_cabecalho = pd.DataFrame({
        "Parâmetro": [
            "Data da Análise",
            "CNPJ do Sacado",
            "Rating Atribuído",
            "Taxa Selic + Prêmio (a.a.)",
            "Prazo Ajustado (dias)",
        ],
        "Valor": [
            data_hora,
            formatar_cnpj(cnpj),
            classificacao.rating,
            f"{resultado.taxa_total * 100:.2f}%",
            resultado.prazo_dias,
        ],
    })

    df_financeiro = pd.DataFrame({
        "Métrica": [
            "Valor de Face (Bruto)",
            "Valor Presente (Desembolso)",
            "Receita Bruta (Desconto)",
            "Perda Esperada (ECL)",
            "Lucro Econômico (RAROC)",
            "Margem Real (%)",
        ],
        "Valor (R$)": [
            valor,
            resultado.valor_presente,
            resultado.desconto_bruto,
            resultado.perda_esperada,
            resultado.lucro_raroc,
            resultado.margem_real / 100,  # Será formatado como percentual
        ],
    })

    df_risco = pd.DataFrame({
        "Parâmetro": [
            "PD Anual (Probability of Default)",
            "LGD (Loss Given Default)",
            "Prêmio de Risco do Rating (a.a.)",
        ],
        "Valor": [
            classificacao.pd_anual,
            0.50,
            classificacao.premio_anual,
        ],
    })

    # =========================================================================
    # Escrita do Excel
    # =========================================================================
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_cabecalho.to_excel(writer, sheet_name="Parecer_Gestao", index=False, startrow=2)
        df_financeiro.to_excel(writer, sheet_name="Parecer_Gestao", index=False, startrow=11)
        df_risco.to_excel(writer, sheet_name="Parecer_Gestao", index=False, startrow=21)

        wb = writer.book
        ws = writer.sheets["Parecer_Gestao"]

        # Título principal
        ws["A1"] = "FIDC INSIGHT — Parecer do Comitê de Crédito"
        ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="0A0A0B")
        ws.merge_cells("A1:B1")
        ws.row_dimensions[1].height = 28

        # Títulos de seção
        _titulo_secao(ws, "A3", "IDENTIFICAÇÃO DA OPERAÇÃO")
        _titulo_secao(ws, "A12", "ANÁLISE FINANCEIRA (RAROC)")
        _titulo_secao(ws, "A22", "PARÂMETROS DE RISCO")

        # Estiliza cabeçalhos das tabelas
        for row in (4, 13, 23):
            for col in ("A", "B"):
                cell = ws[f"{col}{row}"]
                cell.fill = NAVY_FILL
                cell.font = WHITE_BOLD
                cell.alignment = CENTER
                cell.border = BORDER

        # Linhas zebradas + bordas nas tabelas
        for linha_inicio, linha_fim in [(5, 9), (14, 19), (24, 26)]:
            for r in range(linha_inicio, linha_fim + 1):
                fill = ALT_FILL if r % 2 == 0 else None
                for c in ("A", "B"):
                    cell = ws[f"{c}{r}"]
                    cell.border = BORDER
                    if fill:
                        cell.fill = fill
                    cell.font = DEFAULT_FONT
                    cell.alignment = LEFT

        # Formatação numérica da tabela financeira
        for r in range(14, 19):  # Valores R$
            ws[f"B{r}"].number_format = 'R$ #,##0.00'
        ws["B19"].number_format = "0.00%"  # Margem real

        # Formatação de percentuais na tabela de risco
        for r in range(24, 27):
            ws[f"B{r}"].number_format = "0.00%"

        # Larguras das colunas
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 28

        # Nota de rodapé
        ultima_linha = 29
        ws[f"A{ultima_linha}"] = (
            "Documento gerado pelo sistema FIDC Insight · "
            f"Emitido em {data_hora}"
        )
        ws[f"A{ultima_linha}"].font = Font(
            name="Calibri", size=9, italic=True, color="64748B"
        )
        ws.merge_cells(f"A{ultima_linha}:B{ultima_linha}")

        # Congela o título
        ws.freeze_panes = "A4"

    return output.getvalue()


def _titulo_secao(ws, cell_ref: str, texto: str) -> None:
    """Aplica o estilo de título de seção."""
    ws[cell_ref] = texto
    ws[cell_ref].font = Font(name="Calibri", size=12, bold=True, color="8B6A1A")
    col = cell_ref[0]
    linha = cell_ref[1:]
    ws.merge_cells(f"{col}{linha}:B{linha}")
